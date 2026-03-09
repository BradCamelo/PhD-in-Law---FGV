"""

=========================================================================
Coleta automática de dados via APIs públicas:

  ① IBGE API Localidades  → lista dos 223 municípios da PB (códigos IBGE)
       https://servicodados.ibge.gov.br/api/v1/localidades/estados/PB/municipios

  ② IBGE SIDRA Tabela 9514 → população Censo 2022 por município
       https://apisidra.ibge.gov.br/values/t/9514/n3/25/n6/all/v/93/p/2022

  ③ SICONFI/FINBRA API     → receita corrente e despesa corrente 2022
       https://apidatalake.tesouro.gov.br/ords/siconfi/tt/rgf
       (Relatório de Gestão Fiscal — RCL; RREO — Balanço Orçamentário)

  ④ IPEADATA API           → IDHM 2010 por município
       http://www.ipeadata.gov.br/api/odata4/ValoresSerie(SERCODIGO='ADH_IDHM')

  ⑤ OSRM API (OpenStreetMap) → distância rodoviária até João Pessoa
       http://router.project-osrm.org/route/v1/driving/


Notas:
  - Municípios sem retorno na API fiscal recebem valor NaN (transparente).
  - A coleta do OSRM é feita coordenada a coordenada para evitar bloqueios por excesso de requisições.

"""

import time
import requests
import pandas as pd
import numpy as np
from tqdm import tqdm
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ══════════════════════════════════════════════════════════════════════════════
# BLOCO 1 — Lista de municípios da Paraíba (IBGE Localidades API)
# ══════════════════════════════════════════════════════════════════════════════
print("=" * 60)
print("BLOCO 1 — Municípios da PB via IBGE Localidades API")
print("=" * 60)

URL_LOCALIDADES = "https://servicodados.ibge.gov.br/api/v1/localidades/estados/PB/municipios"

resp = requests.get(URL_LOCALIDADES, timeout=30)
resp.raise_for_status()
muns_raw = resp.json()

df_muns = pd.DataFrame([
    {"cod_ibge": str(m["id"]), "Município": m["nome"]}
    for m in muns_raw
]).sort_values("Município").reset_index(drop=True)

print(f"  → {len(df_muns)} municípios encontrados")


# ══════════════════════════════════════════════════════════════════════════════
# BLOCO 2 — População Censo 2022 (IBGE SIDRA Tabela 9514)
# ══════════════════════════════════════════════════════════════════════════════
print("\n" + "=" * 60)
print("BLOCO 2 — População Censo 2022 via IBGE SIDRA (tabela 9514)")
print("=" * 60)

# Tabela 9514 — Censo 2022; variável 93 = população residente; N6 = município
# N3[25] = UF Paraíba
URL_SIDRA = "https://apisidra.ibge.gov.br/values/t/9514/n3/25/n6/all/v/93/p/2022"
    

resp2 = requests.get(URL_SIDRA, timeout=60)
resp2.raise_for_status()
pop_json = resp2.json()

# Primeiro item é o cabeçalho

pop_rows = pop_json[1:] if isinstance(pop_json, list) and len(pop_json) > 1 else pop_json

def parse_int_br(x):
    if x in ("-", "...", None, ""):
        return np.nan
    x = str(x).replace(".", "").replace(",", ".")
    try:
        return int(float(x))
    except ValueError:
        return np.nan

registros = []
for r in pop_rows:
    # Dependendo da estrutura retornada, o código do município pode vir em D2C
    # (quando D1 é UF e D2 é município) ou em D1C.
    cod_mun = r.get("D2C") or r.get("D1C")
    valor = r.get("V")

    if cod_mun:
        registros.append({
            "cod_ibge_7": str(cod_mun)[:7],
            "pop_2022": parse_int_br(valor)
        })

df_pop = pd.DataFrame(registros).drop_duplicates(subset=["cod_ibge_7"])

df_muns["cod_ibge_7"] = df_muns["cod_ibge"].str[:7]

df = df_muns.merge(
    df_pop[["cod_ibge_7", "pop_2022"]],
    on="cod_ibge_7",
    how="left"
)

print(f"  → Pop. 2022 preenchida em {df['pop_2022'].notna().sum()} / {len(df)} municípios")


# ══════════════════════════════════════════════════════════════════════════════
# BLOCO 3 — Receita Corrente e Despesa Corrente 2022 (SICONFI API — FINBRA)
# ══════════════════════════════════════════════════════════════════════════════
print("\n" + "=" * 60)
print("BLOCO 3 — Finanças Municipais 2022 via SICONFI/FINBRA API")
print("=" * 60)

# Endpoint: Contas Anuais (DCA) — Demonstrativo I-C (Receitas) e I-D (Despesas)
# an_exercicio=2022 | co_tipo_demonstrativo=DCA | an_referencia=2022
# Filtra por UF=PB (co_uf=25), esfera municipal (tp_esfera=M)
BASE_SICONFI = "https://apidatalake.tesouro.gov.br/ords/siconfi/tt"

params_receita = {
    "an_exercicio":          2022,
    "in_periodicidade":      "A",          # Anual
    "nr_periodo":            1,
    "co_tipo_demonstrativo": "DCA",
    "no_anexo":              "DCA-Anexo I-C",   # Receitas Orçamentárias
    "co_esfera":             "M",          # Municipal
    "co_uf":                 "PB",
    "limit":                 10000,
}

params_despesa = {
    "an_exercicio":          2022,
    "in_periodicidade":      "A",
    "nr_periodo":            1,
    "co_tipo_demonstrativo": "DCA",
    "no_anexo":              "DCA-Anexo I-D",   # Despesas Orçamentárias
    "co_esfera":             "M",
    "co_uf":                 "PB",
    "limit":                 10000,
}

def fetch_siconfi(endpoint, params):
    url = f"{BASE_SICONFI}/{endpoint}"
    resp = requests.get(url, params=params, timeout=60)
    resp.raise_for_status()
    data = resp.json()
    return pd.DataFrame(data.get("items", data))

print("  Buscando receitas correntes (DCA Anexo I-C) …")
try:
    df_rec_raw = fetch_siconfi("rreo", {
        "an_exercicio":          2022,
        "in_periodicidade":      "A",
        "nr_periodo":            6,           # 6º bimestre = acumulado anual
        "co_tipo_demonstrativo": "RREO",
        "no_anexo":              "RREO-Anexo 01",
        "co_esfera":             "M",
        "co_uf":                 "PB",
        "limit":                 50000,
    })
    # Filtrar linha de Receitas Correntes
    mask_rec = df_rec_raw["coluna"].str.contains("Receitas Correntes", na=False) & \
               (df_rec_raw["rotulo"].str.strip() == "RECEITAS CORRENTES")
    df_rec = (df_rec_raw[mask_rec]
              .groupby("co_ibge")["valor"]
              .sum()
              .reset_index()
              .rename(columns={"valor": "receita_corrente_rs"}))
    print(f"  → Receitas: {len(df_rec)} registros")
except Exception as e:
    print(f"  ⚠ RREO não disponível ({e}); tentando DCA …")
    df_rec = pd.DataFrame(columns=["co_ibge", "receita_corrente_rs"])

# Fallback: DCA (Contas Anuais)
if df_rec.empty:
    try:
        df_dca = fetch_siconfi("dca", params_receita)
        mask = df_dca["conta"].str.startswith("1.0", na=False)  # Receitas Correntes
        df_rec = (df_dca[mask]
                  .groupby("co_ibge")["valor"]
                  .sum()
                  .reset_index()
                  .rename(columns={"valor": "receita_corrente_rs"}))
        print(f"  → DCA receitas: {len(df_rec)} registros")
    except Exception as e2:
        print(f"  ⚠ DCA também falhou ({e2}); receitas ficarão como NaN")
        df_rec = pd.DataFrame(columns=["co_ibge", "receita_corrente_rs"])

print("  Buscando despesas correntes (DCA Anexo I-D) …")
try:
    df_desp_raw = fetch_siconfi("rreo", {
        "an_exercicio":          2022,
        "in_periodicidade":      "A",
        "nr_periodo":            6,
        "co_tipo_demonstrativo": "RREO",
        "no_anexo":              "RREO-Anexo 01",
        "co_esfera":             "M",
        "co_uf":                 "PB",
        "limit":                 50000,
    })
    mask_desp = df_desp_raw["coluna"].str.contains("Despesas Correntes", na=False)
    df_desp = (df_desp_raw[mask_desp]
               .groupby("co_ibge")["valor"]
               .sum()
               .reset_index()
               .rename(columns={"valor": "despesa_corrente_rs"}))
    print(f"  → Despesas: {len(df_desp)} registros")
except Exception as e:
    print(f"  ⚠ Despesas RREO não disponível ({e}); tentando DCA …")
    df_desp = pd.DataFrame(columns=["co_ibge", "despesa_corrente_rs"])

if df_desp.empty:
    try:
        df_dca_d = fetch_siconfi("dca", params_despesa)
        mask = df_dca_d["conta"].str.startswith("3.0", na=False)  # Despesas Correntes
        df_desp = (df_dca_d[mask]
                   .groupby("co_ibge")["valor"]
                   .sum()
                   .reset_index()
                   .rename(columns={"valor": "despesa_corrente_rs"}))
        print(f"  → DCA despesas: {len(df_desp)} registros")
    except Exception as e2:
        print(f"  ⚠ DCA despesas falhou ({e2}); despesas ficarão como NaN")
        df_desp = pd.DataFrame(columns=["co_ibge", "despesa_corrente_rs"])

# Merge finanças
df["co_ibge_6"] = df["cod_ibge"].str[:6]
if not df_rec.empty:
    df_rec["co_ibge_6"] = df_rec["co_ibge"].astype(str).str[:6]
    df = df.merge(df_rec[["co_ibge_6", "receita_corrente_rs"]],
                  on="co_ibge_6", how="left")
    df["receita_corrente_rs"] = df["receita_corrente_rs"] / 1000  # → R$ mil
else:
    df["receita_corrente_rs"] = np.nan

if not df_desp.empty:
    df_desp["co_ibge_6"] = df_desp["co_ibge"].astype(str).str[:6]
    df = df.merge(df_desp[["co_ibge_6", "despesa_corrente_rs"]],
                  on="co_ibge_6", how="left")
    df["despesa_corrente_rs"] = df["despesa_corrente_rs"] / 1000
else:
    df["despesa_corrente_rs"] = np.nan

print(f"  → Receita preenchida: {df['receita_corrente_rs'].notna().sum()} municípios")
print(f"  → Despesa preenchida: {df['despesa_corrente_rs'].notna().sum()} municípios")


# ══════════════════════════════════════════════════════════════════════════════
# BLOCO 4 — IDHM 2010 (IPEADATA API)
# ══════════════════════════════════════════════════════════════════════════════
print("\n" + "=" * 60)
print("BLOCO 4 — IDHM 2010 via IPEADATA API")
print("=" * 60)

URL_IPEA = (
    "http://www.ipeadata.gov.br/api/odata4/"
    "ValoresSerie(SERCODIGO='ADH_IDHM')"
    "?$filter=NIVNOME eq 'Municípios' and VALVALOR ne null"
    "&$select=TERCODIGO,VALVALOR,VALDATA"
    "&$top=10000"
)

try:
    resp4 = requests.get(URL_IPEA, timeout=60)
    resp4.raise_for_status()
    ipea_data = resp4.json()["value"]
    df_idhm_raw = pd.DataFrame(ipea_data)
    # Filtrar ano 2010
    df_idhm_raw["ano"] = pd.to_datetime(df_idhm_raw["VALDATA"]).dt.year
    df_idhm = (df_idhm_raw[df_idhm_raw["ano"] == 2010]
               [["TERCODIGO", "VALVALOR"]]
               .rename(columns={"TERCODIGO": "cod_ibge_6_ipea", "VALVALOR": "IDHM"}))
    # Código do IPEADATA tem 6 dígitos (sem dígito verificador)
    df_idhm["cod_ibge_6_ipea"] = df_idhm["cod_ibge_6_ipea"].astype(str).str.zfill(6)
    df["co_ibge_6"] = df["cod_ibge"].str[:6]
    df = df.merge(df_idhm, left_on="co_ibge_6",
                  right_on="cod_ibge_6_ipea", how="left")
    print(f"  → IDHM preenchido: {df['IDHM'].notna().sum()} municípios")
except Exception as e:
    print(f"  ⚠ IPEADATA falhou ({e}); IDHM ficará como NaN")
    df["IDHM"] = np.nan


# ══════════════════════════════════════════════════════════════════════════════
# BLOCO 5 — Coordenadas + Distância rodoviária até João Pessoa (OSRM API)
# ══════════════════════════════════════════════════════════════════════════════

print("\n" + "=" * 60)
print("BLOCO 5 — Coordenadas e distâncias via IBGE + OSRM")
print("=" * 60)

# 5a. Coordenadas dos municípios via IBGE Malha (centróides)
URL_COORD = (
    "https://servicodados.ibge.gov.br/api/v3/malhas/estados/25"
    "/municipios?formato=application/json&qualidade=minima"
)

# Buscar coordenadas pela API de localidades
# usando o endpoint de geometria simplificada
def get_centroid(cod_ibge7: str) -> tuple:
    """Retorna (lon, lat) do centróide do município via IBGE Malhas API."""
    url = (
        f"https://servicodados.ibge.gov.br/api/v3/malhas/municipios/{cod_ibge7}"
        "?formato=application/json&qualidade=minima"
    )
    try:
        r = requests.get(url, timeout=15)
        r.raise_for_status()
        geo = r.json()
        coords = geo["features"][0]["geometry"]["coordinates"]
        # coords pode ser Polygon ou MultiPolygon — calcular centróide simples
        if geo["features"][0]["geometry"]["type"] == "Polygon":
            pts = coords[0]
        else:  # MultiPolygon
            pts = coords[0][0]
        lon = sum(p[0] for p in pts) / len(pts)
        lat = sum(p[1] for p in pts) / len(pts)
        return (lon, lat)
    except Exception:
        return (None, None)

# João Pessoa: -34.8641, -7.1153 (lon, lat)
JP_LON, JP_LAT = -34.8641, -7.1153

def road_distance_km(lon: float, lat: float) -> float:
    """Distância rodoviária (km) entre ponto e João Pessoa via OSRM."""
    if lon is None or lat is None:
        return np.nan
    url = (
        f"http://router.project-osrm.org/route/v1/driving/"
        f"{lon},{lat};{JP_LON},{JP_LAT}"
        "?overview=false"
    )
    try:
        r = requests.get(url, timeout=15)
        r.raise_for_status()
        data = r.json()
        dist_m = data["routes"][0]["distance"]
        return round(dist_m / 1000, 1)
    except Exception:
        return np.nan

print("  Coletando centróides e calculando distâncias rodoviárias …")
print("  (pode demorar ~5 minutos para 223 municípios)\n")

lons, lats, dists = [], [], []

for _, row in tqdm(df.iterrows(), total=len(df), desc="  Municípios"):
    lon, lat = get_centroid(row["cod_ibge_7"])
    lons.append(lon)
    lats.append(lat)
    dist = road_distance_km(lon, lat)
    dists.append(dist)
    time.sleep(0.25)   # respeitar rate limit das APIs públicas

df["longitude"] = lons
df["latitude"]  = lats
df["dist_km"]   = dists

n_dist = df["dist_km"].notna().sum()
print(f"\n  → Distâncias calculadas: {n_dist} / {len(df)} municípios")


# ══════════════════════════════════════════════════════════════════════════════
# BLOCO 6 — Organizar DataFrame final
# ══════════════════════════════════════════════════════════════════════════════
print("\n" + "=" * 60)
print("BLOCO 6 — DataFrame final")
print("=" * 60)

df_final = df[[
    "Município",
    "cod_ibge",
    "pop_2022",
    "receita_corrente_rs",
    "despesa_corrente_rs",
    "dist_km",
    "IDHM",
]].copy()

df_final.columns = [
    "Município",
    "Código IBGE",
    "População (2022)",
    "Receita Corrente (R$ mil)",
    "Despesa Corrente (R$ mil)",
    "Distância p/ Capital (km)",
    "IDHM (2010)",
]

df_final = df_final.sort_values("Município").reset_index(drop=True)
df_final.index += 1

# Relatório de cobertura
total  = len(df_final)
cob_pop  = df_final["População (2022)"].notna().sum()
cob_rec  = df_final["Receita Corrente (R$ mil)"].notna().sum()
cob_desp = df_final["Despesa Corrente (R$ mil)"].notna().sum()
cob_dist = df_final["Distância p/ Capital (km)"].notna().sum()
cob_idh  = df_final["IDHM (2010)"].notna().sum()

print(f"\n  Municípios        : {total}")
print(f"  População         : {cob_pop}/{total} ({cob_pop/total*100:.1f}%)")
print(f"  Receita Corrente  : {cob_rec}/{total} ({cob_rec/total*100:.1f}%)")
print(f"  Despesa Corrente  : {cob_desp}/{total} ({cob_desp/total*100:.1f}%)")
print(f"  Distância         : {cob_dist}/{total} ({cob_dist/total*100:.1f}%)")
print(f"  IDHM              : {cob_idh}/{total} ({cob_idh/total*100:.1f}%)")

print("\nPrimeiras linhas:")
print(df_final.head(8).to_string())

print("\nEstatísticas descritivas:")
print(df_final.describe().round(2).to_string())


# ══════════════════════════════════════════════════════════════════════════════
# BLOCO 7 — Exportar Excel formatado
# ══════════════════════════════════════════════════════════════════════════════
print("\n" + "=" * 60)
print("BLOCO 7 — Exportando Excel")
print("=" * 60)

wb = Workbook()

h_fill  = PatternFill("solid", start_color="1B4F72")
alt_f   = PatternFill("solid", start_color="D6EAF8")
white   = PatternFill("solid", start_color="FFFFFF")
total_f = PatternFill("solid", start_color="1A5276")
src_f   = PatternFill("solid", start_color="F0F3F4")
thin    = Side(style="thin", color="BFC9CA")
brd     = Border(left=thin, right=thin, top=thin, bottom=thin)

ws = wb.active
ws.title = "Municípios da Paraíba"

# Título
ws.merge_cells("A1:H1")
t = ws["A1"]
t.value = "MUNICÍPIOS DA PARAÍBA — INDICADORES SOCIOECONÔMICOS"
t.font  = Font(name="Arial", bold=True, size=13, color="FFFFFF")
t.fill  = PatternFill("solid", start_color="0A2E4A")
t.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 28

# Subtítulo com fontes
ws.merge_cells("A2:H2")
src = ws["A2"]
src.value = (
    "Fontes: IBGE Localidades API | IBGE SIDRA tab.9514 (Pop. 2022) | "
    "SICONFI/FINBRA API (Rec./Desp. 2022) | IPEADATA (IDHM 2010) | OSRM/OSM (Distâncias)"
)
src.font  = Font(name="Arial", italic=True, size=8, color="555555")
src.fill  = src_f
src.alignment = Alignment(horizontal="center")

# Cabeçalho
headers = [
    "Município", "Código\nIBGE", "População\n(2022)",
    "Receita Corrente\n(R$ mil)", "Despesa Corrente\n(R$ mil)",
    "Distância p/\nCapital (km)", "IDHM\n(2010)", "Cobertura\nDados"
]
for c, h in enumerate(headers, 1):
    cell = ws.cell(row=3, column=c, value=h)
    cell.font      = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    cell.fill      = h_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = brd
ws.row_dimensions[3].height = 32

# Dados
for row_i, row in df_final.iterrows():
    excel_r = row_i + 2
    fill    = alt_f if row_i % 2 == 0 else white

    # Cobertura: conta quantas variáveis estão preenchidas
    cob = sum([
        pd.notna(row["População (2022)"]),
        pd.notna(row["Receita Corrente (R$ mil)"]),
        pd.notna(row["Despesa Corrente (R$ mil)"]),
        pd.notna(row["Distância p/ Capital (km)"]),
        pd.notna(row["IDHM (2010)"]),
    ])
    cob_str = f"{cob}/5"

    vals = [
        row["Município"], row["Código IBGE"], row["População (2022)"],
        row["Receita Corrente (R$ mil)"], row["Despesa Corrente (R$ mil)"],
        row["Distância p/ Capital (km)"], row["IDHM (2010)"], cob_str
    ]
    fmts  = [None, None, "#,##0", "#,##0", "#,##0", "#,##0", "0.000", None]
    aligns = ["left","center","right","right","right","center","center","center"]

    for c, (v, fmt, aln) in enumerate(zip(vals, fmts, aligns), 1):
        cell = ws.cell(row=excel_r, column=c, value=v if pd.notna(v) else None)
        cell.font      = Font(name="Arial", size=9)
        cell.fill      = fill
        cell.border    = brd
        cell.alignment = Alignment(horizontal=aln)
        if fmt and pd.notna(v):
            cell.number_format = fmt
        # Destacar células NaN em amarelo
        if v != cob_str and pd.isna(v):
            cell.fill = PatternFill("solid", start_color="FFF9C4")

# Totais / médias
total_row = len(df_final) + 4
ws.cell(row=total_row, column=1, value="TOTAL / MÉDIA").font = Font(
    name="Arial", bold=True, size=9, color="FFFFFF"
)
ws.cell(row=total_row, column=1).fill      = total_f
ws.cell(row=total_row, column=1).border    = brd
ws.cell(row=total_row, column=1).alignment = Alignment(horizontal="center")

for c in range(2, 9):
    cell       = ws.cell(row=total_row, column=c)
    cell.fill  = total_f
    cell.font  = Font(name="Arial", bold=True, size=9, color="FFFFFF")
    cell.border= brd
    cl = get_column_letter(c)
    if c == 3:   # população = soma
        cell.value = f"=SUM({cl}4:{cl}{total_row-1})"
        cell.number_format = "#,##0"
        cell.alignment = Alignment(horizontal="right")
    elif c in [4, 5]:  # receita/despesa = soma
        cell.value = f"=SUM({cl}4:{cl}{total_row-1})"
        cell.number_format = "#,##0"
        cell.alignment = Alignment(horizontal="right")
    elif c in [6, 7]:  # distância/IDHM = média
        cell.value = f"=AVERAGE({cl}4:{cl}{total_row-1})"
        cell.number_format = "0.0" if c == 6 else "0.000"
        cell.alignment = Alignment(horizontal="center")
    else:
        cell.value = ""

# Larguras
col_widths = [32, 10, 13, 18, 18, 14, 9, 10]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A4"

# ── Aba: Fontes e Metadados ────────────────────────────────────────────────
ws_meta = wb.create_sheet("Fontes e Metadados")
ws_meta.column_dimensions["A"].width = 90

meta_rows = [
    ("FONTES DE DADOS E METADADOS",                                            "title"),
    ("",                                                                       ""),
    ("① LISTA DE MUNICÍPIOS",                                                  "h2"),
    ("API: IBGE Localidades",                                                  "key"),
    ("URL: https://servicodados.ibge.gov.br/api/v1/localidades/estados/PB/municipios", "val"),
    ("Retorno: JSON com nome e código IBGE de todos os 223 municípios da PB.", "body"),
    ("",                                                                       ""),
    ("② POPULAÇÃO — CENSO 2022",                                               "h2"),
    ("API: IBGE SIDRA — Tabela 9514",                                          "key"),
    ("URL: https://apisidra.ibge.gov.br/values/t/9514/n6/in N3 25/v/93/p/2022","val"),
    ("Variável 93 = População residente. Nível territorial N6 = Município.",   "body"),
    ("",                                                                       ""),
    ("③ RECEITA E DESPESA CORRENTE — 2022",                                    "h2"),
    ("API: SICONFI / FINBRA — Tesouro Nacional",                               "key"),
    ("URL base: https://apidatalake.tesouro.gov.br/ords/siconfi/tt/",          "val"),
    ("Demonstrativo: RREO Anexo 01 (6º bimestre = acumulado anual 2022).",     "body"),
    ("Receita Corrente: conta com código iniciado em 1. | Despesa: código 3.", "body"),
    ("Valores em R$ (convertidos para R$ mil no script).",                     "body"),
    ("",                                                                       ""),
    ("④ IDHM — ÍNDICE DE DESENVOLVIMENTO HUMANO MUNICIPAL 2010",               "h2"),
    ("API: IPEADATA — Série ADH_IDHM",                                         "key"),
    ("URL: http://www.ipeadata.gov.br/api/odata4/ValoresSerie(SERCODIGO='ADH_IDHM')","val"),
    ("Fonte primária: Atlas do Desenvolvimento Humano — PNUD/IPEA/FJP, 2013.", "body"),
    ("Ano de referência: Censo 2010 (última edição completa disponível).",     "body"),
    ("",                                                                       ""),
    ("⑤ DISTÂNCIAS RODOVIÁRIAS ATÉ JOÃO PESSOA",                              "h2"),
    ("API: OSRM — Open Source Routing Machine (OpenStreetMap)",                "key"),
    ("URL: http://router.project-osrm.org/route/v1/driving/{lon},{lat};{JP_lon},{JP_lat}","val"),
    ("Centróides dos municípios: IBGE Malhas API v3 (geometria simplificada).", "body"),
    ("Referência: Coordenadas de João Pessoa = -34.8641° lon, -7.1153° lat.",  "body"),
    ("",                                                                       ""),
    ("LEGENDA — COBERTURA DE DADOS",                                           "h2"),
    ("Células amarelas = dado não disponível na API para aquele município.",   "body"),
    ("Coluna 'Cobertura Dados': n/5 indica quantas das 5 variáveis estão preenchidas.", "body"),
]

for r, (txt, kind) in enumerate(meta_rows, 1):
    cell = ws_meta.cell(row=r, column=1, value=txt)
    if kind == "title":
        cell.font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color="0A2E4A")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws_meta.row_dimensions[r].height = 24
    elif kind == "h2":
        cell.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
        cell.fill = h_fill
        ws_meta.row_dimensions[r].height = 18
    elif kind == "key":
        cell.font = Font(name="Arial", bold=True, size=10)
    elif kind == "val":
        cell.font = Font(name="Arial", size=9, color="1A5276", italic=True)
    else:
        cell.font = Font(name="Arial", size=10)
        cell.alignment = Alignment(wrap_text=True)

output = "municipios_paraiba_APIs.xlsx"
wb.save(output)
print(f"\n  Arquivo salvo: {output}")
print("  Concluído!")
