"""
Randomização Estratificada com Emparelhamento Pairwise
==================================================================

"""

import pandas as pd
import numpy as np
import networkx as nx
from scipy.optimize import linear_sum_assignment
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings("ignore")

np.random.seed(1981)

# ── 1. Dados ──────────────────────────────────────────────────────────────────

df = pd.read_csv("municipios_paraiba_final_v2.csv", encoding="utf-8")

# Renomear as colunas para manter o padrão do script inicial
df = df.rename(columns={
    "Municipio": "Município",
    "Populacao": "Pop",
    "Receita_Corrente": "Receita",
    "Despesa_Corrente": "Despesa",
    "Distancia_Capital": "Dist",
    "IDHM": "IDHM"
})

# Ordenar alfabeticamente
df = df.sort_values("Município").reset_index(drop=True)

# Número de municípios
N = len(df)




# ── 2. Estratificação por quartis da distância ────────────────────────────────
df["Estrato_Dist"] = pd.qcut(
    df["Dist"], q=4,
    labels=["Q1_Próximo", "Q2", "Q3", "Q4_Distante"]
)

print("\nDistribuição por estrato de distância:")
print(df["Estrato_Dist"].value_counts().sort_index())

# ── 3. Distância de Mahalanobis ───────────────────────────────────────────────
# Variáveis de matching com transformação log para Receita e Despesa
df["log_Receita"] = np.log(df["Receita"])
df["log_Despesa"] = np.log(df["Despesa"])
match_cols = ["log_Receita", "log_Despesa", "IDHM"]

# Matriz de covariância global (captura correlação entre as variáveis)
X       = df[match_cols].values
cov_inv = np.linalg.inv(np.cov(X.T))

def mahal_dist(u, v, VI):
    """Distância de Mahalanobis entre dois vetores u e v."""
    diff = u - v
    return float(np.sqrt(diff @ VI @ diff))

# ── 4. Emparelhamento dentro de cada estrato (Algoritmo Húngaro) ─────────────
all_pairs = []

for stratum in sorted(df["Estrato_Dist"].unique(), key=str):
    sub  = df[df["Estrato_Dist"] == stratum]
    idx  = sub.index.tolist()
    n    = len(idx)
    
    if n < 2: continue

    if n % 2 != 0:
        print(f"Estrato {stratum} tem número ímpar ({n}). Ajuste necessário.")
        continue

    Xsub = df.loc[idx, match_cols].values

G = nx.Graph()
for a in range(n):
        for b in range(a + 1, n):
            d = mahal_dist(Xsub[a], Xsub[b], cov_inv)
            G.add_edge(idx[a], idx[b], weight=d)

matching = nx.algorithms.matching.min_weight_matching(
        G, weight="weight"
    )

for u, v in matching:
        d = G[u][v]["weight"]
        all_pairs.append((u, v, d))

print(f"Estrato {stratum}: {n} municípios → {len(matching)} pares")

# ── 5. Selecionar os 43 melhores pares ────────────────────────────────────────
# Ordenar por distância de Mahalanobis crescente (melhor qualidade = menor dist)
all_pairs.sort(key=lambda x: x[2])
selected = all_pairs[:43]

print(f"\nTotal de pares gerados: {len(all_pairs)}")
print(f"Pares selecionados (43 menores distâncias): {len(selected)}")

# ── 6. Aleatorização dentro de cada par (sorteio binário) ─────────────────────
grupo_a, grupo_b = [], []

for pair_id, (i, j, dist_m) in enumerate(selected, 1):
    if np.random.randint(0, 2) == 0:
        ga, gb = i, j
    else:
        ga, gb = j, i
    grupo_a.append({"Par": pair_id, "idx": ga, "Dist_Mahal": round(dist_m, 4)})
    grupo_b.append({"Par": pair_id, "idx": gb, "Dist_Mahal": round(dist_m, 4)})

# ── 7. Montar DataFrames dos grupos ──────────────────────────────────────────
def build_group(group_list, label):
    rows = []
    for g in group_list:
        r = df.loc[g["idx"]]
        rows.append({
            "Par":                        g["Par"],
            "Grupo":                      label,
            "Município":                  r["Município"],
            "Estrato_Distância":          str(r["Estrato_Dist"]),
            "População (2022)":           r["Pop"],
            "Receita Corrente (R$ mil)":  r["Receita"],
            "Despesa Corrente (R$ mil)":  r["Despesa"],
            "Distância p/ Capital (km)":  r["Dist"],
            "IDHM":                       r["IDHM"],
            "Dist_Mahalanobis":           g["Dist_Mahal"],
        })
    return pd.DataFrame(rows)

df_a = build_group(grupo_a, "A")
df_b = build_group(grupo_b, "B")

# Municípios não selecionados
sel_set = {g["idx"] for g in grupo_a} | {g["idx"] for g in grupo_b}
df_ns   = df.loc[[i for i in range(N) if i not in sel_set]].copy()

# ── 8. Verificação do balanço entre grupos ────────────────────────────────────
print("\n── VERIFICAÇÃO DO BALANÇO ──")
print(f"{'Variável':<12} {'Grupo A':>12} {'Grupo B':>12} {'Dif. %':>8}")
print("-" * 48)
for col_df, col_g in [("Pop","População (2022)"), ("Receita","Receita Corrente (R$ mil)"),
                       ("Despesa","Despesa Corrente (R$ mil)"), ("Dist","Distância p/ Capital (km)"),
                       ("IDHM","IDHM")]:
    ma = df_a[col_g].mean()
    mb = df_b[col_g].mean()
    d  = abs(ma - mb) / ma * 100
    print(f"{col_df:<12} {ma:>12.2f} {mb:>12.2f} {d:>7.1f}%")

# ── 9. Exportar Excel formatado ───────────────────────────────────────────────
wb = Workbook()

# Paleta de estilos
h_fill  = PatternFill("solid", start_color="0A2E4A")
a_fill  = PatternFill("solid", start_color="1A5276")
b_fill  = PatternFill("solid", start_color="145A32")
alt_a   = PatternFill("solid", start_color="D6EAF8")
alt_b   = PatternFill("solid", start_color="D5F5E3")
alt_ns  = PatternFill("solid", start_color="FDEBD0")
white   = PatternFill("solid", start_color="FFFFFF")
thin    = Side(style="thin", color="BFC9CA")
brd     = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr(cell, fill, txt, size=10, bold=True):
    cell.value     = txt
    cell.font      = Font(name="Arial", bold=bold, size=size, color="FFFFFF")
    cell.fill      = fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = brd

def dat(cell, fill, val, fmt=None, align="center"):
    cell.value     = val
    cell.font      = Font(name="Arial", size=9)
    cell.fill      = fill
    cell.alignment = Alignment(horizontal=align)
    cell.border    = brd
    if fmt:
        cell.number_format = fmt

# ── Aba 1: Pares lado a lado ──────────────────────────────────────────────────
ws1 = wb.active
ws1.title = "Pares Randomizados"

ws1.merge_cells("A1:N1")
hdr(ws1["A1"], h_fill,
    "RANDOMIZAÇÃO ESTRATIFICADA COM EMPARELHAMENTO PAIRWISE — DISTÂNCIA DE MAHALANOBIS", size=12)
ws1.row_dimensions[1].height = 28

# Sub-header de grupos
for rng, txt, fill in [("A2:A2","Par",h_fill), ("B2:G2","◼  GRUPO A  ◼",a_fill),
                        ("H2:M2","◼  GRUPO B  ◼",b_fill), ("N2:N2","Dist. Mahal.",h_fill)]:
    ws1.merge_cells(rng)
    hdr(ws1[rng.split(":")[0]], fill, txt, size=10)
ws1.row_dimensions[2].height = 18

# Cabeçalho das colunas
col_h = ["Par",
         "Município","Estrato","Pop.","Receita\n(R$mil)","Despesa\n(R$mil)","IDHM",
         "Município","Estrato","Pop.","Receita\n(R$mil)","Despesa\n(R$mil)","IDHM",
         "Dist.\nMahal."]
for c, h in enumerate(col_h, 1):
    fill = a_fill if 2 <= c <= 7 else (b_fill if 8 <= c <= 13 else h_fill)
    hdr(ws1.cell(row=3, column=c), fill, h, size=9)
ws1.row_dimensions[3].height = 30

# Dados dos pares
for row_i, (ra, rb) in enumerate(zip(df_a.itertuples(), df_b.itertuples()), 4):
    alt = row_i % 2 == 0
    fa  = PatternFill("solid", start_color="D6EAF8") if alt else white
    fb  = PatternFill("solid", start_color="D5F5E3") if alt else white

    vals = [
        (ra.Par,                          h_fill, None,    "center"),
        (ra.Município,                    fa,     None,    "left"),
        (ra.Estrato_Distância,            fa,     None,    "center"),
        (ra._5,                           fa,     "#,##0", "right"),
        (ra._6,                           fa,     "#,##0", "right"),
        (ra._7,                           fa,     "#,##0", "right"),
        (ra.IDHM,                         fa,     "0.000", "center"),
        (rb.Município,                    fb,     None,    "left"),
        (rb.Estrato_Distância,            fb,     None,    "center"),
        (rb._5,                           fb,     "#,##0", "right"),
        (rb._6,                           fb,     "#,##0", "right"),
        (rb._7,                           fb,     "#,##0", "right"),
        (rb.IDHM,                         fb,     "0.000", "center"),
        (ra.Dist_Mahalanobis,             fa,     "0.0000","center"),
    ]
    for c, (v, f, fmt, aln) in enumerate(vals, 1):
        cell = ws1.cell(row=row_i, column=c)
        if c == 1:
            cell.value = v
            cell.font  = Font(name="Arial", bold=True, size=9, color="FFFFFF")
            cell.fill  = h_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = brd
        else:
            dat(cell, f, v, fmt, aln)

# Linha de médias
mr = len(df_a) + 4
for c in range(1, 15):
    cell = ws1.cell(row=mr, column=c)
    fill = a_fill if 1 <= c <= 7 else (b_fill if 8 <= c <= 13 else h_fill)
    cell.fill  = fill
    cell.border = brd
    cell.font  = Font(name="Arial", bold=True, size=9, color="FFFFFF")
    if c == 1:
        cell.value = "MÉDIA"
        cell.alignment = Alignment(horizontal="center")
    elif c in [4,5,6,10,11,12]:
        cl = get_column_letter(c)
        cell.value = f"=AVERAGE({cl}4:{cl}{mr-1})"
        cell.number_format = "#,##0"
        cell.alignment = Alignment(horizontal="right")
    elif c in [7,13]:
        cl = get_column_letter(c)
        cell.value = f"=AVERAGE({cl}4:{cl}{mr-1})"
        cell.number_format = "0.000"
        cell.alignment = Alignment(horizontal="center")
    else:
        cell.value = ""

# Larguras de coluna
for c, w in enumerate([5,28,16,10,14,14,8,28,16,10,14,14,8,11], 1):
    ws1.column_dimensions[get_column_letter(c)].width = w
ws1.freeze_panes = "B4"

# ── Aba 2: Grupo A ────────────────────────────────────────────────────────────
def write_group_sheet(wb, df_g, title, sheet_title, fill_h, fill_alt):
    ws = wb.create_sheet(sheet_title)
    ws.merge_cells("A1:I1")
    hdr(ws["A1"], fill_h, title, size=12)
    ws.row_dimensions[1].height = 25

    hdrs = ["Par","Município","Estrato Dist.","Pop. (2022)","Receita Corrente\n(R$ mil)",
            "Despesa Corrente\n(R$ mil)","Dist. Capital\n(km)","IDHM","Dist. Mahal."]
    for c, h in enumerate(hdrs, 1):
        hdr(ws.cell(row=2, column=c), fill_h, h, size=9)
    ws.row_dimensions[2].height = 30

    for r, row in enumerate(df_g.itertuples(), 3):
        f = fill_alt if r % 2 == 0 else white
        vals = [row.Par, row.Município, row.Estrato_Distância,
                row._5, row._6, row._7, row._8, row.IDHM, row.Dist_Mahalanobis]
        fmts = [None, None, None, "#,##0","#,##0","#,##0","#,##0","0.000","0.0000"]
        alns = ["center","left","center","right","right","right","center","center","center"]
        for c, (v, fmt, aln) in enumerate(zip(vals, fmts, alns), 1):
            dat(ws.cell(row=r, column=c), f, v, fmt, aln)

    for c, w in enumerate([5,28,16,12,15,15,10,8,11], 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "B3"

write_group_sheet(wb, df_a, "GRUPO A — 43 MUNICÍPIOS RANDOMIZADOS", "Grupo A",
                  a_fill, alt_a)
write_group_sheet(wb, df_b, "GRUPO B — 43 MUNICÍPIOS RANDOMIZADOS", "Grupo B",
                  b_fill, alt_b)

# ── Aba 4: Não selecionados ────────────────────────────────────────────────────
ws4 = wb.create_sheet("Não Selecionados")
ns_h = PatternFill("solid", start_color="784212")
ws4.merge_cells("A1:G1")
hdr(ws4["A1"], PatternFill("solid", start_color="6E2F00"),
    f"MUNICÍPIOS NÃO SELECIONADOS ({len(df_ns)} municípios)", size=12)
ws4.row_dimensions[1].height = 25

for c, h in enumerate(["Município","Pop. (2022)","Receita (R$ mil)","Despesa (R$ mil)",
                        "Dist. Capital (km)","IDHM","Estrato Dist."], 1):
    hdr(ws4.cell(row=2, column=c), ns_h, h, size=9)
ws4.row_dimensions[2].height = 30

for r, row in enumerate(df_ns.sort_values("Município").itertuples(), 3):
    f = alt_ns if r % 2 == 0 else white
    vals = [row.Município, row.Pop, row.Receita, row.Despesa,
            row.Dist, row.IDHM, str(row.Estrato_Dist)]
    fmts = [None,"#,##0","#,##0","#,##0","#,##0","0.000",None]
    alns = ["left","right","right","right","center","center","center"]
    for c, (v, fmt, aln) in enumerate(zip(vals, fmts, alns), 1):
        dat(ws4.cell(row=r, column=c), f, v, fmt, aln)

for c, w in enumerate([30,12,14,14,12,8,16], 1):
    ws4.column_dimensions[get_column_letter(c)].width = w
ws4.freeze_panes = "A3"

# ── Aba 5: Metodologia ────────────────────────────────────────────────────────
ws5 = wb.create_sheet("Metodologia")
ws5.column_dimensions["A"].width = 90

meto = [
    ("METODOLOGIA DA RANDOMIZAÇÃO",                                                    "title"),
    ("",                                                                               ""),
    ("1. BASE DE DADOS E UNIVERSO DE ANÁLISE",                                         "h2"),
    ("O procedimento utiliza dados dos municípios do Estado da Paraíba contendo as seguintes variáveis:", "body"),
    ("população residente (Censo IBGE 2022), receita corrente municipal, despesa corrente municipal,", "body"),
    ("distância rodoviária até a capital João Pessoa e Índice de Desenvolvimento Humano Municipal (IDHM).", "body"),
    ("As variáveis fiscais foram obtidas da base FINBRA da Secretaria do Tesouro Nacional.", "body"),
    ("O IDHM foi extraído do Atlas do Desenvolvimento Humano no Brasil (PNUD).",       "body"),
    ("",                                                                               ""),
    ("2. ESTRATIFICAÇÃO GEOGRÁFICA",                                                   "h2"),
    ("Os municípios foram inicialmente divididos em quatro estratos definidos pelos quartis", "body"),
    ("da distribuição da distância rodoviária até João Pessoa.",                       "body"),
    ("Essa estratificação reduz heterogeneidade territorial e assegura comparabilidade", "body"),
    ("entre unidades experimentais com níveis semelhantes de proximidade à capital.",  "body"),
    ("",                                                                               ""),
    ("3. VARIÁVEIS DE EMPARELHAMENTO",                                                 "h2"),
    ("Dentro de cada estrato os municípios são comparados com base em três variáveis estruturais:", "body"),
    ("  • log da Receita Corrente municipal",                                          "body"),
    ("  • log da Despesa Corrente municipal",                                          "body"),
    ("  • Índice de Desenvolvimento Humano Municipal (IDHM)",                          "body"),
    ("A transformação logarítmica reduz a assimetria das variáveis fiscais e mitiga",  "body"),
    ("a influência de municípios de grande porte.",                                    "body"),
    ("",                                                                               ""),
    ("4. DISTÂNCIA DE MAHALANOBIS",                                                    "h2"),
    ("A similaridade entre municípios é medida pela distância de Mahalanobis:",       "body"),
    ("d_M(u,v) = sqrt[(u − v)' · Σ⁻¹ · (u − v)]",                                      "body"),
    ("onde u e v representam os vetores de características dos municípios e Σ⁻¹ é",    "body"),
    ("a inversa da matriz de covariância das variáveis utilizadas no emparelhamento.", "body"),
    ("Essa métrica considera simultaneamente escala e correlação entre variáveis,",    "body"),
    ("produzindo uma medida estatisticamente mais adequada de proximidade.",           "body"),
    ("",                                                                               ""),
    ("5. EMPARELHAMENTO ÓTIMO",                                                        "h2"),
    ("Dentro de cada estrato constrói-se um grafo não direcionado em que cada",        "body"),
    ("município corresponde a um vértice e cada aresta representa um possível",        "body"),
    ("emparelhamento entre dois municípios.",                                          "body"),
    ("O peso de cada aresta corresponde à distância de Mahalanobis entre eles.",       "body"),
    ("O algoritmo de minimum-weight matching seleciona o conjunto de pares",           "body"),
    ("que minimiza a soma total das distâncias dentro de cada estrato,",                "body"),
    ("produzindo um emparelhamento globalmente ótimo.",                                "body"),
    ("",                                                                               ""),
    ("6. SELEÇÃO DOS PARES",                                                           "h2"),
    ("Todos os pares gerados nos estratos são ordenados de acordo com a distância",    "body"),
    ("de Mahalanobis. Os pares com menor distância representam municípios",            "body"),
    ("mais semelhantes em termos socioeconômicos e fiscais.",                          "body"),
    ("A amostra final do experimento é composta pelos 43 pares com menor distância,",  "body"),
    ("totalizando 86 municípios selecionados para participação no experimento.",       "body"),
    ("",                                                                               ""),
    ("7. RANDOMIZAÇÃO DENTRO DO PAR",                                                  "h2"),
    ("Após o emparelhamento, a alocação ao tratamento é realizada por sorteio",         "body"),
    ("binário dentro de cada par de municípios.",                                      "body"),
    ("Em cada par, um município é designado para o Grupo A (tratamento) e o outro",    "body"),
    ("para o Grupo B (controle).",                                                     "body"),
    ("A randomização utiliza semente pseudoaleatória fixa (np.random.seed(1981)),",      "body"),
    ("garantindo reprodutibilidade integral do procedimento.",                         "body"),
    ("",                                                                               ""),
    ("8. FONTES DE DADOS",                                                             "h2"),
    ("  • População: IBGE, Censo Demográfico 2022",                                    "body"),
    ("  • Receita e Despesa: FINBRA / Secretaria do Tesouro Nacional",                 "body"),
    ("  • IDHM: Atlas do Desenvolvimento Humano no Brasil (PNUD)",                     "body"),
    ("  • Distâncias rodoviárias: cálculo geográfico a partir de João Pessoa (PB)",    "body"),
    ("",                                                                               ""),
    ("Implementação computacional: Python 3 — pandas, numpy e networkx.",              "body"),
]

for r, (txt, kind) in enumerate(meto, 1):
    cell = ws5.cell(row=r, column=1, value=txt)
    if kind == "title":
        cell.font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color="0A2E4A")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws5.row_dimensions[r].height = 22
    elif kind == "h2":
        cell.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
        cell.fill = a_fill
        ws5.row_dimensions[r].height = 18
    else:
        cell.font = Font(name="Arial", size=10)
        cell.alignment = Alignment(wrap_text=True)

output = "randomizacao_mahalanobis_PB.xlsx"
wb.save(output)
print(f"\nArquivo salvo: {output}")
